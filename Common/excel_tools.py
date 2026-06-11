import logging
import re
from Common.read_and_save_tool import ConfigTools
import openpyxl
import os
from Common import logger
import json

logger = logger.logger


class ExcelTools:
    def __init__(self):
        self.file_path = self._get_file_path()
        self.workbook = None
        self.sheet = None
        self._load_workbook()
        self.config = ConfigTools()

    def _get_file_path(self):
        """获取Data文件下的excel路径"""
        return os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Data', 'api_testcase.xlsx')

    def _load_workbook(self):
        """加载Excel工作簿"""
        try:
            if not hasattr(self, 'file_path') or not self.file_path:
                self.file_path = self._get_file_path()
            if os.path.exists(self.file_path):
                self.workbook = openpyxl.load_workbook(self.file_path)
            else:
                raise FileNotFoundError(f"指定的文件不存在: {self.file_path}")
            self.sheet = self.workbook.active
        except Exception as e:
            logger.error(f"无法加载Excel文件 '{self.file_path}': {str(e)}")
            raise IOError(f"无法加载Excel文件 '{self.file_path}': {str(e)}")

    def get_sheet(self, sheet_name):
        """
        根据sheet名称获取sheet对象
        :param sheet_name: 工作表名称
        :return: sheet对象
        :raises ValueError: 当工作表不存在时
        """
        if not sheet_name:
            raise ValueError("工作表名称不能为空")

        if not self.workbook:
            self._load_workbook()

        if sheet_name not in self.workbook.sheetnames:
            available_sheets = ", ".join(self.workbook.sheetnames)
            raise ValueError(f"工作表 '{sheet_name}' 未找到。可用的工作表: {available_sheets}")

        return self.workbook[sheet_name]

    def get_row_data_by_test_case_name(self, sheet_name, test_case_name):
        """
        根据用例名称获取该行的所有数据
        :param test_case_name: 用例名称
        :param sheet_name: 可选，指定的工作表名称
        :return: 包含该行所有数据的字典（键为列标题，值为对应单元格的值）
        :raises ValueError: 当未找到用例名称时
        """
        if not sheet_name or not test_case_name:
            raise ValueError("sheet_name 和 test_case_name 不能为空")

        target_sheet = self.get_sheet(sheet_name)

        if not target_sheet:
            raise RuntimeError("未设置有效的工作表")

        headers = [cell.value for cell in target_sheet[1]]

        for row in target_sheet.iter_rows(min_row=2):
            if any(cell.value == test_case_name for cell in row):
                row_data = {}
                for col_idx, cell_value in enumerate(row):
                    if col_idx < len(headers):
                        row_data[headers[col_idx]] = cell_value.value
                return row_data

        raise ValueError(f"未找到用例名称 '{test_case_name}'")

    def get_case_data(self, sheet_name, test_case_name):
        """
        获取用例数据
        :param test_case_name: 用例名称
        :param sheet_name: 工作表名称
        :return: 包含用例数据的字典
        """
        try:
            case_data = self.get_row_data_by_test_case_name(sheet_name, test_case_name)
            if not isinstance(case_data, dict):
                raise ValueError("获取的测试用例数据不是字典类型")

            keys_mapping = {
                "用例 ID": "用例 ID",
                "用例名称": "用例名称",
                "请求方法": "请求方法",
                "接口路径": "接口路径",
                "优先级": "优先级",
                "前置条件": "前置条件",
                "请求体（JSON）": "请求体（JSON）",
                "预期状态码": "预期状态码",
                "预计消费额度": "预计消费额度",
                "响应时间阈值 (ms)": "响应时间阈值 (ms)",
                "执行结果": "用例结果",
            }

            test_case_data = {}
            for key, mapped_key in keys_mapping.items():
                test_case_data[mapped_key] = case_data.get(key)

            return test_case_data
        except Exception as e:
            logging.error(f"读取测试用例数据失败: {e}")
            raise

    def write_result(self, sheet_name, test_case_name, header_titles, result_data):
        """
        写入测试结果到Excel
        :param sheet_name: 工作表名称
        :param test_case_name: 测试用例名称
        :param header_titles: 结果列标题
        :param result_data: 测试结果数据
        :return: None
        """
        try:
            # 获取工作表
            target_sheet = self.get_sheet(sheet_name)

            # 查找用例所在的行
            headers = [cell.value for cell in target_sheet[1]]
            result_column_index = None

            # 找到结果列的索引
            for i, header in enumerate(headers):
                if header == header_titles:
                    result_column_index = i + 1  # Excel列索引从1开始
                    break

            if result_column_index is None:
                raise ValueError("未找到结果列")

            # 查找用例名称所在的行
            target_row = None
            for row_idx, row in enumerate(target_sheet.iter_rows(min_row=2), start=2):
                for cell in row:
                    if cell.value == test_case_name:
                        target_row = row_idx
                        break
                if target_row:
                    break

            if target_row is None:
                raise ValueError(f"未找到用例 '{test_case_name}'")

            # 写入结果
            target_sheet.cell(row=target_row, column=result_column_index, value=result_data)

            # 保存文件
            self.workbook.save(self.file_path)

        except Exception as e:
            logging.error(f"写入测试结果失败: {e}")
            raise

    def process_url_placeholder(self, url_template, replace_values):
        """
        处理URL模板中的占位符：先提取占位符，再替换为指定值

        参数：
            url_template: 包含占位符的URL模板（如 "https://example.com/{id}"）
            replace_values: 替换值字典（如 {"id": "123"}）

        返回：
            替换后的完整URL字符串
        """
        # 步骤1：提取所有占位符（{...}中的内容）
        pattern = r"\{(.*?)\}"
        placeholders = re.findall(pattern, url_template)

        # 检查替换值是否覆盖所有占位符
        for placeholder in placeholders:
            if placeholder not in replace_values:
                logger.error(f"缺少替换值：URL中的占位符 '{placeholder}' 未在replace_values中找到")
                raise ValueError(f"缺少替换值：URL中的占位符 '{placeholder}' 未在replace_values中找到")

        # 步骤2：替换所有占位符
        processed_url = url_template
        for placeholder, value in replace_values.items():
            # 确保占位符带{}（如将"id"转为"{id}"）
            placeholder_with_braces = f"{{{placeholder}}}"
            processed_url = processed_url.replace(placeholder_with_braces, str(value))
        logger.info(f"URL已处理: {processed_url}")

        return processed_url

    def get_url(self, path, ping_data: str = None, replace_data: dict = None, dict_data: dict = None):
        """
        获取测试用例的URL方法（GET/POST等）

        参数：
            path: 接口路径
            ping_data: 查询参数字符串
            replace_data: URL占位符替换数据
            dict_data: 字典形式的查询参数

        返回：
            完整URL字符串
        """
        try:
            if not path:
                raise ValueError("路径不能为空")

            authority = self.config.get_url_data()

            # 根据不同参数类型处理URL
            if replace_data and isinstance(replace_data, dict):
                url_template = authority + path
                return self.process_url_placeholder(url_template, replace_data)
            elif ping_data:
                return f"{authority}{path}?{ping_data}"
            elif dict_data and isinstance(dict_data, dict):
                query_params = []
                for key, value in dict_data.items():
                    if value is None:
                        continue
                    elif isinstance(value, list):
                        for item in value:
                            if item is not None:
                                query_params.append(f"{key}[]={item}")
                    else:
                        query_params.append(f"{key}={value}")

                if query_params:
                    return f"{authority}{path}?{'&'.join(query_params)}"
                else:
                    return f"{authority}{path}"
            else:
                return authority + path

        except Exception as e:
            logger.error(f"获取URL方法失败: {e}")
            raise

    # def _replace_variables(self, data, variables):
    #     """
    #     替换请求参数中的变量（如 ${user_id} → 12345）
    #     :param data: 原始请求参数（dict/list/str）
    #     :param variables: 变量字典（如 {"user_id": 12345, "token": "xxx"}）
    #     :return: 替换后的参数
    #     """
    #     try:
    #         # 处理 None 或空数据
    #         if not data:
    #             return data
    #
    #         # 处理字典类型
    #         if isinstance(data, dict):
    #             result = {}
    #             for key, value in data.items():
    #                 # 处理键中的变量
    #                 new_key = self._replace_string_variables(key, variables) if isinstance(key, str) else key
    #                 # 处理值中的变量
    #                 new_value = self._replace_variables(value, variables)
    #                 result[new_key] = new_value
    #             logger.debug(f"替换后的参数: {result}")
    #             return result
    #
    #         # 处理列表类型
    #         elif isinstance(data, list):
    #             return [self._replace_variables(item, variables) for item in data]
    #
    #         # 处理字符串类型
    #         elif isinstance(data, str):
    #             return self._replace_string_variables(data, variables)
    #
    #         # 其他类型（int, float, bool等）直接返回
    #         else:
    #             return data
    #
    #     except Exception as e:
    #         logger.error(f"替换变量失败，数据类型: {type(data)}, 数据值: {data}, 错误: {e}")
    #         raise

    #
    # def _replace_string_variables(self, text, variables):
    #     """
    #     替换请求体内的参数
    #     替换单个字符串中的所有变量占位符
    #     :param text: 原始字符串
    #     :param variables: 变量字典
    #     :return: 替换后的字符串
    #     """
    #     if not isinstance(text, str):
    #         return text
    #
    #     result = text
    #     # 查找所有 ${...} 格式的变量
    #     pattern = r'\$\{([^}]+)\}'
    #     matches = re.findall(pattern, text)
    #
    #     for var_name in matches:
    #         if var_name in variables:
    #             var_value = variables[var_name]
    #
    #             # 检查变量是否在引号内（带引号的变量）
    #             quoted_pattern = f'"\$\{{{var_name}\}}"'
    #             unquoted_pattern = f'\$\{{{var_name}\}}'
    #
    #             # 如果变量在引号内，只需替换值部分，不添加额外引号
    #             if quoted_pattern in result:
    #                 # 直接替换整个带引号的部分
    #                 result = result.replace(quoted_pattern, f'"{var_value}"')
    #             # 如果变量不在引号内
    #             elif unquoted_pattern in result:
    #                 # 根据变量类型决定是否添加引号
    #                 if isinstance(var_value, (int, float, bool)):
    #                     if isinstance(var_value, bool):
    #                         replacement = str(var_value).lower()
    #                     else:
    #                         replacement = str(var_value)
    #                 else:
    #                     replacement = f'"{var_value}"'
    #                 result = result.replace(unquoted_pattern, replacement)
    #         else:
    #             logger.warning(f"变量 {var_name} 未找到，保留原始占位符")
    #
    #     return result

    def update_test_case_result(self, sheet_name,
                                test_case_name,
                                variables=None,
                                ping_data: str = None,
                                replace_data: str = None,
                                dict_data: dict = None):
        """
        更新测试用例结果并返回测试所需数据

        Args:
            sheet_name (str): Excel工作表名称
            test_case_name (str): 测试用例名称
            variables (dict, optional): 变量替换字典，默认为空字典
            ping_data (str, optional): 查询参数字符串
            replace_data (str, optional): URL占位符替换数据
            dict_data (dict, optional): 字典形式的查询参数

        Returns:
            tuple: (method, url, data, assert_code, assert_amount, case_id)
                - method: 请求方法
                - url: 完整请求URL
                - data: 处理后的请求体数据
                - assert_code: 预期状态码
                - assert_amount: 预计消费额度
                - case_id: 用例ID

        Raises:
            ValueError: 当必需参数缺失或数据格式错误时
            Exception: 其他处理异常
        """

        if variables is None:
            variables = {}

        try:
            test_case_data = self.get_case_data(sheet_name, test_case_name)
            if not test_case_data:
                raise ValueError(f"未找到测试用例数据: {sheet_name} - {test_case_name}")

            case_id = test_case_data.get('用例ID', None)
            path = test_case_data.get('接口路径', None)
            method = test_case_data.get('请求方法', None)
            assert_code = test_case_data.get('预期状态码', None)
            # assert_amount = test_case_data.get('预计消费额度', None)

            if not path:
                raise ValueError("接口路径不能为空")

            url = self.get_url(path, ping_data=ping_data, replace_data=replace_data, dict_data=dict_data)

            body_str = test_case_data.get('请求体（JSON）')

            # 解析请求体数据
            request_body = {}
            if isinstance(body_str, str) and body_str.strip():
                try:
                    request_body = json.loads(body_str)
                except json.JSONDecodeError as e:
                    logger.error(f"JSON解析失败: {body_str}")
                    raise ValueError(f"请求体JSON格式错误: {e}")
            elif isinstance(body_str, dict):
                request_body = body_str.copy()  # 创建副本避免修改原始数据

            # 使用字典update方法合并变量
            if isinstance(request_body, dict) and isinstance(variables, dict):
                updated_body = request_body.copy()  # 创建副本
                updated_body.update(variables)  # 合并变量
                data = json.dumps(updated_body, ensure_ascii=False)
            elif request_body:
                data = json.dumps(request_body, ensure_ascii=False)
            else:
                data = '{}'  # 空对象
            logger.info(f"更新测试用例结果: {sheet_name} - {test_case_name} - {data}")
            return method, url, data, assert_code, case_id

        except Exception as e:
            logger.error(f"更新测试用例结果失败: {e}")
            raise

    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.sheet = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


if __name__ == '__main__':
    from Common.read_and_save_tool import ConfigTools
    import hashlib
    data = {'type': 'bank', 'payment_type': 'local', 'currency': 'SG', 'bank_address': {'country': 'SGD'}, 'bank_name': '中国农业银行', 'account_no_iban': '7154062325813894', 'bic_swift': 'ABOCCNBJ', 'payee_type': 'personal', 'relation': 'GRAND_FATHER', 'payee_address': {'postCode': '765234', 'country': 'SGD', 'state': 'Central Singapore', 'city': 'Berlin', 'line1': 'Park Avenue 93'}, 'first_name': 'Thomas', 'last_name': 'Johnson'}
    b ={"payee_name":"rayn","currency":"USDT","chain_id":"42161","wallet_address":"156"}
    try:
        config = ConfigTools()
        tools = ExcelTools()

        password = config.get_login_data('PASSWORD')
        email = config.get_login_data('EMAIL')

        if not password:
            raise ValueError("PASSWORD配置未找到或为空")
        if not email:
            raise ValueError("EMAIL配置未找到或为空")

        variables = {
            'email': email,
            'password': hashlib.md5(password.encode('utf-8')).hexdigest()
        }
        body ={
            "wallet_id": "e56f09e4-e2cc-4453-a513-3b90d565fa18",
            "chain_id": "42161",
            "payee_id": "abe2a694-f069-4ce1-a791-34913b149872",
            "amount": 0.1}
        # try:
        #     tools.update_test_case_result('Wallte_page', '钱包-加密资产转出', variables=body)
        # except Exception as e:
        #     print(f"处理测试用例结果时出错: {e}")
        #
        # try:
        #     tools.update_test_case_result('Login_page', '登陆成功', variables=variables)
        # except Exception as e:
        #     print(f"处理测试用例结果时出错: {e}")
        # try:
        #     tools.update_test_case_result('Wallte_page', '钱包-创建法币收款地址', variables=data)
        # except Exception as e:
        #     print(f"处理测试用例结果时出错: {e}")
        try:
            tools.update_test_case_result('Wallte_page', '钱包-创建加密收款地址', variables=b)
        except Exception as e:
            print(f"处理测试用例结果时出错: {e}")
    except Exception as e:
        logger.error(f"测试执行失败: {e}")
        raise
    finally:
        if 'tools' in locals():
            tools.close()

